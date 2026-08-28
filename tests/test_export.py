import csv
import io

import pytest
from httpx import ASGITransport, AsyncClient

from tests.test_api import _prepare_app


@pytest.fixture()
async def client(tmp_path):
    app = _prepare_app(tmp_path / "export.db")
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        yield c
    await app.state.router.shutdown()


async def _seed(client, user_id="u_export"):
    headers = {"X-Api-Key": "test-api-key"}
    for mid, content in [
        ("e1", "今天微信吃饭花了25元"),
        ("e2", "昨天支付宝打车花了35元"),
        ("e3", "今天银行卡收入100元"),
    ]:
        r = await client.post(
            "/api/v1/message/inbound",
            headers=headers,
            json={"platform": "qq", "user_id": user_id, "chat_id": "c_x", "message_id": mid, "content": content},
        )
        assert r.status_code == 200


async def test_export_csv_matches_db(client):
    await client.post(
        "/admin/api/login", json={"username": "admin", "password": "test-admin-pass-123"}
    )
    await _seed(client)
    resp = await client.get("/admin/api/export?user_id=u_export&format=csv")
    assert resp.status_code == 200
    text = resp.content.decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text)))
    assert len(rows) == 3
    by_desc = {r["备注"]: r for r in rows}
    assert by_desc["微信吃饭"]["金额(元)"] == "25.00"
    assert by_desc["微信吃饭"]["类型"] == "支出"
    assert by_desc["微信吃饭"]["账户"] == "微信"
    assert by_desc["支付宝打车" if "支付宝打车" in by_desc else "打车"]["金额(元)"] == "35.00"
    assert by_desc["银行卡收入"]["类型"] == "收入"
    assert by_desc["银行卡收入"]["金额(元)"] == "100.00"


async def test_export_xlsx_matches_db(client):
    from openpyxl import load_workbook

    await client.post(
        "/admin/api/login", json={"username": "admin", "password": "test-admin-pass-123"}
    )
    await _seed(client)
    resp = await client.get("/admin/api/export?user_id=u_export&format=xlsx")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    data = [dict(zip(header, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    assert len(data) == 3
    by_desc = {r["备注"]: r for r in data}
    assert by_desc["微信吃饭"]["金额(元)"] == 25.0
    assert by_desc["银行卡收入"]["类型"] == "收入"


async def test_export_date_range(client):
    await client.post(
        "/admin/api/login", json={"username": "admin", "password": "test-admin-pass-123"}
    )
    headers = {"X-Api-Key": "test-api-key"}
    for mid, content in [
        ("r1", "2026年8月20日微信吃饭花了25元"),
        ("r2", "2026年8月21日支付宝打车花了35元"),
        ("r3", "今天银行卡收入100元"),
    ]:
        r = await client.post(
            "/api/v1/message/inbound", headers=headers,
            json={"platform": "qq", "user_id": "u_export", "chat_id": "c_x", "message_id": mid, "content": content},
        )
        assert r.status_code == 200
    resp = await client.get("/admin/api/export?user_id=u_export&format=csv&start=2026-08-20&end=2026-08-20")
    rows = list(csv.DictReader(io.StringIO(resp.content.decode("utf-8-sig"))))
    assert len(rows) == 1
    assert rows[0]["日期"] == "2026-08-20"
