import csv
import datetime
import io
import re
from typing import Any, Optional

from paas.models import CategoryRow, ImportResult, ParsedItem
from paas.modules.account.parser import match_category

# 优先列：同角色多列时优先；兜底列：主流记账 App 的差异化表头
HEADER_PREFERRED: dict[str, list[str]] = {
    "date": ["日期", "时间", "消费日期", "记账日期", "交易日期", "date", "time"],
    "amount": ["金额", "消费金额", "支出金额", "花费", "费用", "amount", "price", "money"],
    "category": ["分类", "类别", "类目", "category"],
    "description": ["备注", "描述", "说明", "明细", "用途", "摘要", "description", "note"],
    "type": ["类型", "收支", "收/支", "收支类型", "交易类型", "type"],
    "account": ["账户", "账号", "account"],
}
HEADER_FALLBACK: dict[str, list[str]] = {
    "date": ["交易时间"],
    "amount": [],
    "category": ["交易分类", "消费类别"],
    "description": ["商品", "商品说明", "消费内容", "项目"],
    "type": [],
    "account": ["支付方式", "收/付款方式", "钱包"],
}

SUPPORTED_SUFFIXES = (".csv", ".xlsx", ".xls")


def _normalize_account(value: Any) -> str:
    """把"微信支付/余额宝/储蓄卡"等说法归一化为标准账户名。"""
    from paas.modules.account.parser import detect_accounts

    text = str(value).strip()
    if not text:
        return ""
    accs = detect_accounts(text)
    return accs[0] if accs else text


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value).strip().lower())


def _detect_columns(header_row: list[Any]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for aliases in (HEADER_PREFERRED, HEADER_FALLBACK):
        for idx, cell in enumerate(header_row):
            key = _norm_header(cell)
            for role, role_aliases in aliases.items():
                if role in cols:
                    continue
                if any(key == _norm_header(a) for a in role_aliases):
                    cols[role] = idx
                    break
    return cols


def _find_header(rows: list[list[Any]]) -> tuple[dict[str, int], int]:
    for i, row in enumerate(rows[:10]):
        cols = _detect_columns(row)
        if len(cols) >= 2 and "amount" in cols:
            return cols, i
    return {"amount": 0}, 0


def _parse_amount_cell(value: Any) -> Optional[int]:
    """返回金额（分）；负数/收入返回 None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        num = float(value)
    else:
        s = str(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
        m = re.search(r"(-?\d+(?:\.\d+)?)", s)
        if not m:
            return None
        num = float(m.group(1))
    if num <= 0:
        return None
    return int(round(num * 100))


def _parse_date_cell(value: Any, base_year: int) -> Optional[datetime.date]:
    if value is None or str(value).strip() in ("", "-", "/"):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value).strip()
    # 兼容带时间的长格式，如 "2026-08-28 16:34:12" / "2026/08/28 16:34"
    m = re.search(r"(\d{4})\s*[-/.年]\s*(\d{1,2})\s*[-/.月]\s*(\d{1,2})", s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%Y%m%d"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.fullmatch(r"(\d{1,2})[月/\-.](\d{1,2})[日号]?", s)
    if m:
        try:
            return datetime.date(base_year, int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def _rows_from_csv(data: bytes) -> list[list[str]]:
    text = None
    for enc in ("utf-8-sig", "gbk", "utf-8"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = data.decode("utf-8", errors="replace")
    return [list(row) for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(data: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        rows.append(["" if v is None else v for v in row])
    wb.close()
    return rows


def parse_import(
    data: bytes,
    filename: str,
    categories: list[CategoryRow],
) -> tuple[ImportResult, list[ParsedItem]]:
    result = ImportResult()
    items: list[ParsedItem] = []
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = _rows_from_csv(data)
    elif lower.endswith((".xlsx", ".xls")):
        rows = _rows_from_xlsx(data)
    else:
        result.errors.append("不支持的文件类型，请发送 .csv / .xlsx 文件")
        return result, items

    cols, header_idx = _find_header(rows)
    data_rows = rows[header_idx + 1:]
    base_year = datetime.date.today().year

    for row_idx, row in enumerate(data_rows, start=1):
        result.total_rows += 1
        if not any(str(c).strip() for c in row):
            continue
        raw_amount = (
            row[cols["amount"]]
            if cols.get("amount") is not None and cols["amount"] < len(row)
            else None
        )
        amount_cents = _parse_amount_cell(raw_amount)
        tx_type = "expense"
        if amount_cents is None and str(raw_amount).strip().startswith("-"):
            # 负数金额（无类型列时）视为收入
            m = re.search(r"-?\d+(?:\.\d+)?", str(raw_amount))
            if m:
                amount_cents = int(round(abs(float(m.group())) * 100))
                tx_type = "income"
        if amount_cents is None:
            result.errors.append(f"第{row_idx}行：金额无效")
            continue
        d = _parse_date_cell(
            row[cols["date"]] if cols.get("date") is not None and cols["date"] < len(row) else None,
            base_year,
        )
        if d is None:
            result.errors.append(f"第{row_idx}行：日期无效")
            continue
        raw_cat = row[cols["category"]] if cols.get("category") is not None and cols["category"] < len(row) else ""
        cat = match_category(str(raw_cat), categories) if str(raw_cat).strip() else next(
            (c for c in categories if c.name == "其他"), categories[-1]
        )
        raw_desc = row[cols["description"]] if cols.get("description") is not None and cols["description"] < len(row) else ""
        desc = str(raw_desc).strip() or cat.name
        raw_type = row[cols["type"]] if cols.get("type") is not None and cols["type"] < len(row) else ""
        type_text = str(raw_type).strip()
        if "退款" in type_text:
            tx_type = "refund"
        elif "收入" in type_text or "工资" in type_text:
            tx_type = "income"
        elif "支出" in type_text:
            tx_type = "expense"
        raw_account = row[cols["account"]] if cols.get("account") is not None and cols["account"] < len(row) else ""
        account_name = _normalize_account(raw_account)
        if not account_name and type_text:
            account_name = ""
        items.append(
            ParsedItem(
                expense_date=d,
                category_id=cat.id,
                category_name=cat.name,
                category_icon=cat.icon,
                account_name=account_name,
                tx_type=tx_type,
                amount_cents=amount_cents,
                description=desc,
            )
        )

    result.success_rows = len(items)
    result.failed_rows = result.total_rows - len(items)
    result.errors = result.errors[:50]
    return result, items
